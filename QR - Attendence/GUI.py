import cv2
import numpy as np
import pyzbar.pyzbar as pyzbar
import sys
import time 
import openpyxl
import tkinter as tk
from tkinter import messagebox
import requests

# create or load workbook
try:
    wb = openpyxl.load_workbook("attendance.xlsx")
    sheet = wb.active
except FileNotFoundError:
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Attendance"

# initialize names list with existing data
names = [cell.value for cell in sheet["A"][1:]]

# function to enter data in the attendance file
def enterData(z):
    if z in names:
        pass
    else:
        names.append(z)
        sheet.cell(row=len(names), column=1, value=z)
        # add date and time to the next column
        sheet.cell(row=len(names), column=2, value=time.strftime("%Y-%m-%d %H:%M:%S"))
        # add GPS coordinates to the next column
        lat, lon = getGPS()
        sheet.cell(row=len(names), column=3, value=f"{lat}, {lon}")


# function to check if data is present in attendance file
def checkData(data):
    data = str(data)
    if data in names:
        messagebox.showwarning("Duplicate Entry", "This data is already present in the attendance file.")
    else:
        enterData(data)
        wb.save("attendance.xlsx")  # save the workbook after adding data
        messagebox.showinfo("Attendance", f"Attendance recorded for {data}")

# function to get GPS coordinates
def getGPS():
    url = "http://ip-api.com/json/"
    data = requests.get(url).json()
    lat = data['lat']
    lon = data['lon']
    return lat, lon

# function to start scanning
def startScan():
    cap = cv2.VideoCapture(0)
    while True:
        _, frame = cap.read()
        decoded_objects = pyzbar.decode(frame)
        for obj in decoded_objects:
            checkData(obj.data)
            time.sleep(1)  
            
        cv2.imshow("Scan QR Code", frame)

        if cv2.waitKey(1) & 0xff == ord("s"):
            cv2.destroyAllWindows()
            break

    # release the webcam and save the workbook
    cap.release()
    wb.save("attendance.xlsx")

# function to view attendance
def viewAttendance():
    # read the names, timestamps and GPS coordinates from the worksheet
    names = [cell.value for cell in sheet["A"][1:]]
    timestamps = [cell.value for cell in sheet["B"][1:]]
    gps_coordinates = [cell.value for cell in sheet["C"][1:]]

    # create a list of tuples with name and corresponding timestamp and GPS coordinates
    data = list(zip(names, timestamps, gps_coordinates))

    # create a new tkinter window for displaying attendance
    root = tk.Tk()
    root.title("Attendance")

    # create a tkinter listbox widget and insert the attendance data
    listbox = tk.Listbox(root, width=70)
    for name, timestamp, gps in data:
        listbox.insert(tk.END, f"{name}  -  {timestamp} - {gps}")

    # pack the listbox widget and start the tkinter mainloop
    listbox.pack(padx=10, pady=10)
    root.mainloop()

# create a tkinter window
window = tk.Tk()
window.title("Attendance System")


# create a tkinter label widget
label = tk.Label(window, text="Select an option to continue:", font=("Arial", 16))
label.pack(padx=10, pady=10)

# create tkinter buttons for scan and view attendance
scan_button = tk.Button(window, text="Scan QR Code", font=("Arial", 14), command=startScan)
scan_button.pack(padx=10, pady=10)

view_button = tk.Button(window, text="View Attendance", font=("Arial", 14), command=viewAttendance)
view_button.pack(padx=10, pady=10)

# start the tkinter mainloop
window.mainloop()


