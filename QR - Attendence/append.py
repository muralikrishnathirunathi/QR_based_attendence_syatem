

import cv2
import numpy as np
import pyzbar.pyzbar as pyzbar
import sys
import time 
import openpyxl

# create or load workbook
try:
    wb = openpyxl.load_workbook("attendance.xlsx")
    sheet = wb.active
except FileNotFoundError:
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Attendance"

# initialize names list with existing data
names = [cell.value for cell in sheet["A"][1:]]

# function to enter data in the attendance file
def enterData(z, gps_stamp):
    if z in names:
        pass
    else:
        names.append(z)
        sheet.cell(row=len(names), column=1, value=z)
        # add date, time and GPS stamp to the next columns
        sheet.cell(row=len(names), column=2, value=time.strftime("%Y-%m-%d %H:%M:%S"))
        sheet.cell(row=len(names), column=3, value=gps_stamp)

# function to check if data is present in attendance file
def checkData(data, gps_stamp):
    data = str(data)
    if data in names:
        print("This data is already present in the attendance file.")
    else:
        enterData(data, gps_stamp)
        wb.save("attendance.xlsx")  # save the workbook after adding data
        print(f"Attendance recorded for {data}")

# function to start scanning
def startScan():
    cap = cv2.VideoCapture(0)
    while True:
        _, frame = cap.read()
        decoded_objects = pyzbar.decode(frame)
        for obj in decoded_objects:
            checkData(obj.data, "GPS stamp") # Replace "GPS stamp" with actual GPS data
            time.sleep(1)  
            
        cv2.imshow("Scan QR Code", frame)

        if cv2.waitKey(1) & 0xff == ord("s"):
            cv2.destroyAllWindows()
            break

    # release the webcam and save the workbook
    cap.release()
    wb.save("attendance.xlsx")

# function to view attendance
def viewAttendance():
    # read the names and timestamps from the worksheet
    names = [cell.value for cell in sheet["A"][1:]]
    timestamps = [cell.value for cell in sheet["B"][1:]]
    gps_stamps = [cell.value for cell in sheet["C"][1:]]

    # create a list of tuples with name, timestamp and GPS stamp
    data = list(zip(names, timestamps, gps_stamps))

    # print the attendance data
    print("Attendance:")
    for name, timestamp, gps_stamp in data:
        print(f"{name} - {timestamp} - {gps_stamp}")

# call the functions to start scanning and view attendance
startScan()
viewAttendance()
