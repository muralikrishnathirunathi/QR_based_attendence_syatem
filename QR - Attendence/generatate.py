import qrcode
from openpyxl import load_workbook

# Load Excel sheet
workbook = load_workbook("E:/QR - Attendence/names.xlsx")
sheet = workbook.active

# Loop through IDs and generate QR codes
for row in sheet.iter_rows(min_row=2):
    id = row[0].value  # Access the cell value using index 0
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(id)
    qr.make(fit=True)
    img = qr.make_image(fill_color='black', back_color='white')
    
    # Save QR code as image file
    img.save(f'{id}.png')


